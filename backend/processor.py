import io
import datetime
import unicodedata
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ─── Time helpers ─────────────────────────────────────────────────────────────

def _to_minutes(value: Any) -> int:
    """Convert an openpyxl cell value to total minutes."""
    if value is None or value == "":
        return 0

    if isinstance(value, datetime.time):
        return value.hour * 60 + value.minute

    if isinstance(value, datetime.timedelta):
        return int(value.total_seconds()) // 60

    if isinstance(value, (int, float)):
        if value == 0:
            return 0
        # Excel stores time as fraction of a day; durations > 24 h exceed 1.0
        return round(value * 24 * 60)

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return 0
        parts = s.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1]) if len(parts) > 1 else 0
            return h * 60 + m
        except (ValueError, IndexError):
            return 0

    return 0


def _round_30(total_minutes: int) -> int:
    """30-min rounding rule: remainder < 30 → floor, >= 30 → ceil."""
    if total_minutes <= 0:
        return 0
    h = total_minutes // 60
    m = total_minutes % 60
    return h + 1 if m >= 30 else h


def _to_decimal(minutes: int) -> float:
    return round(minutes / 60, 2)


# ─── Column detection ─────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Lowercase + strip diacritics for fuzzy header matching."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    ).strip()


def _detect_columns(header_row: tuple) -> dict:
    cols: dict = {}
    for i, cell in enumerate(header_row):
        v = _norm(str(cell or ""))
        if v == "empleado":
            cols["empleado"] = i
        elif v == "legajo":
            cols["legajo"] = i
        elif v == "fecha" and "fecha" not in cols:
            cols["fecha"] = i
        elif v == "marcaciones":
            cols["marcaciones"] = i
        elif v == "observaciones":
            cols["observaciones"] = i
        elif v == "normal" and "normal" not in cols:
            cols["normal"] = i
        elif "50" in v and "extra" in v and "ext50" not in cols:
            cols["ext50"] = i
        elif "100" in v and "extra" in v and "ext100" not in cols:
            cols["ext100"] = i
    return cols


# ─── Main processing ──────────────────────────────────────────────────────────

def process_workbook(file_bytes: bytes) -> dict:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Locate header row (search first 15 rows)
    header_idx = -1
    cols: dict = {}
    for i, row in enumerate(rows[:15]):
        cols = _detect_columns(row)
        if "empleado" in cols and "legajo" in cols and "normal" in cols:
            header_idx = i
            break

    if header_idx == -1:
        raise ValueError(
            "No se encontraron columnas obligatorias (Empleado, Legajo, Normal). "
            "Verificá que sea el reporte de marcaciones correcto."
        )

    warnings = []
    if "ext50" not in cols:
        warnings.append("Columna 'Extras 50%' no encontrada — se omite HSEX50")
    if "ext100" not in cols:
        warnings.append("Columna 'Extras 100%' no encontrada — se omite HSEX100")

    employees: dict = {}

    for row in rows[header_idx + 1:]:
        nombre_raw = str(row[cols["empleado"]] or "").strip()

        # Legajo puede venir como float (ej. 268.0) desde Excel
        legajo_val = row[cols["legajo"]]
        if isinstance(legajo_val, float):
            legajo_val = int(legajo_val)
        legajo_raw = str(legajo_val or "").strip()

        if not nombre_raw or not legajo_raw:
            continue
        if _norm(nombre_raw) in ("totales", "total"):
            continue

        if legajo_raw not in employees:
            employees[legajo_raw] = {
                "nombre":      nombre_raw,
                "legajo":      legajo_raw,
                "normal_min":  0,
                "feriado_min": 0,
                "ext50_min":   0,
                "ext100_min":  0,
            }

        normal_min = _to_minutes(row[cols["normal"]])
        ext50_min  = _to_minutes(row[cols["ext50"]])  if "ext50"  in cols else 0
        ext100_min = _to_minutes(row[cols["ext100"]]) if "ext100" in cols else 0

        obs_text = str(row[cols["observaciones"]] or "").upper() if "observaciones" in cols else ""
        mrc_text = str(row[cols["marcaciones"]]   or "").lower() if "marcaciones"   in cols else ""
        is_feriado = "FERIADO" in obs_text or "feriado" in mrc_text

        emp = employees[legajo_raw]
        if is_feriado:
            emp["feriado_min"] += normal_min
        else:
            emp["normal_min"] += normal_min
        emp["ext50_min"]  += ext50_min
        emp["ext100_min"] += ext100_min

    today = datetime.date.today().strftime("%d/%m/%Y")
    tango_rows = []

    for emp in employees.values():
        hsaut   = _to_decimal(emp["normal_min"])
        hsfet   = _to_decimal(emp["feriado_min"])
        hsex50  = _round_30(emp["ext50_min"])
        hsex100 = _round_30(emp["ext100_min"])

        if hsaut   > 0:
            tango_rows.append({"fecha": today, "legajo": emp["legajo"], "codigo": "HSAUT",   "cantidad": hsaut,   "valor": 1, "obs": ""})
        if hsfet   > 0:
            tango_rows.append({"fecha": today, "legajo": emp["legajo"], "codigo": "HSFET",   "cantidad": hsfet,   "valor": 1, "obs": ""})
        if hsex50  > 0:
            tango_rows.append({"fecha": today, "legajo": emp["legajo"], "codigo": "HSEX50",  "cantidad": hsex50,  "valor": 1, "obs": ""})
        if hsex100 > 0:
            tango_rows.append({"fecha": today, "legajo": emp["legajo"], "codigo": "HSEX100", "cantidad": hsex100, "valor": 1, "obs": ""})

    return {
        "rows": tango_rows,
        "warnings": warnings,
        "stats": {
            "total_employees": len(employees),
            "total_rows":      len(tango_rows),
            "with_extras":     sum(1 for e in employees.values() if e["ext50_min"] + e["ext100_min"] > 0),
            "with_feriado":    sum(1 for e in employees.values() if e["feriado_min"] > 0),
        },
    }


# ─── Excel generation ─────────────────────────────────────────────────────────

def generate_excel(tango_rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tango"

    headers = ["Fecha", "Número de legajo", "Código de novedad", "Cantidad", "Valor", "Observaciones"]
    ws.append(headers)

    for r in tango_rows:
        ws.append([r["fecha"], r["legajo"], r["codigo"], r["cantidad"], r["valor"], r["obs"]])

    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, width in enumerate([14, 20, 22, 12, 10, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
